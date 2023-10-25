from openpyxl import *
from openpyxl.styles import Border, Side, Alignment,Font,PatternFill
from Excel.rows_and_cols import *

class Excel():
    def __init__(self,path): # path are bytes not directory names
        #self.path = BytesIO(path)
        self.path = path
        self.error = 0
        # self.rows = {}
        # self.cols = {}
        self.all_object = []
    def openfile(self):
        try:
            self.excel = load_workbook(self.path)
            return self.error
        except FileNotFoundError:
            return self.error
    def allsheets(self):
        return self.excel.sheetnames
    def getrows(self):
        sheet = self.excel[self.excel.sheetnames[0]]
        num_rows = sheet.max_row # จำนวนสมาชิก row สูงสุดใน sheet
        num_cols = sheet.max_column # จำนวนสมาชิก column สูงสุดใน sheet 
        for i in range(1,num_rows+1):
            if i == 1:
                header = each_row(sheet,i,num_rows,num_cols)
            else:
                slave = each_row(sheet,i,num_rows,num_cols)
                self.all_object.append(sub_object_send(str(i-1),header,slave))
        return self.all_object,num_rows-1
    # def getcolumns(self):
    #     sheet = self.excel[self.excel.sheetnames[0]]
    #     num_rows = sheet.max_row # จำนวนสมาชิก row สูงสุดใน sheet
    #     num_cols = sheet.max_column # จำนวนสมาชิก column สูงสุดใน sheet
    #     for i in range(1, num_cols+1):
    #         self.cols.update({"col"+str(i):each_column(sheet,i,num_cols,num_rows)}) #each_column send list datatype and collect in self.cols dict
    #     return self.cols
    def closefile(self):
        self.excel.close()
        
class WriteExcel():
    def __init__(self,save_path):
        self.save_path = save_path
    def create_workbook(self): #เผื่อใช้ในการ sheet ที่ไม่มีการสร้าง File ใหม่
        self.workbook = Workbook()
    def select_sheet(self,sheet_name,position = 0):
        #self.sheet = self.workbook.active
        self.sheet = self.workbook.create_sheet(sheet_name,position)
    def insert_data(self,row_target,col_terget,data,position=None,IsBold = None,fill = None):
        #alignment_style = Alignment(horizontal='left', vertical='center')
        if position is not None:
            if "all" in position:
                if "center" in position:
                    x = y = "center"
                elif "right" in position:
                    x = "right"
                    y = "bottom"
                else:
                    x = "left"
                    y = "top"
            else:
                x,y = position.split(",")
            alignment_style = Alignment(horizontal=x, vertical=y)
            if IsBold == "Bold" or IsBold == "bold":
                self.sheet.cell(row=row_target, column=col_terget, value=data).alignment = alignment_style
                self.sheet.cell(row=row_target, column=col_terget, value=data).font = Font(bold = True)
                if fill is not None:
                    color = PatternFill(start_color="00FFCC99",end_color="00FFCC99",fill_type="solid")
                    self.sheet.cell(row=row_target, column=col_terget, value=data).fill=color
            else:
                self.sheet.cell(row=row_target, column=col_terget, value=data).alignment = alignment_style
                if fill is not None and fill == True:
                    color = PatternFill(start_color="00FFCC99",end_color="00FFCC99",fill_type="solid")
                    self.sheet.cell(row=row_target, column=col_terget, value=data).fill = color
        else:
            if IsBold == "Bold" or IsBold == "bold":
                self.sheet.cell(row=row_target, column=col_terget, value=data).font = Font(bold = True)
                if fill is not None and fill == True:
                    color = PatternFill(start_color="00FFCC99",end_color="00FFCC99",fill_type="solid")
                    self.sheet.cell(row=row_target, column=col_terget, value=data).fill = color
            else:
                self.sheet.cell(row=row_target, column=col_terget, value=data)
                if fill is not None and fill == True:
                    color = PatternFill(start_color="00FFCC99",end_color="00FFCC99",fill_type="solid")
                    self.sheet.cell(row=row_target, column=col_terget, value=data).fill = color
        
        # if position == "center":
        #     alignment_style = Alignment(horizontal='center', vertical='center')
        #     self.sheet.cell(row=row_target, column=col_terget, value=data).alignment = alignment_style
        # else:   
        #     self.sheet.cell(row=row_target, column=col_terget, value=data)
        # if IsBold == "Bold" or IsBold == "bold":
        #     self.sheet.cell(row=row_target, column=col_terget, value=data).font = Font(bold = True)
            
        # self.sheet.cell(row=1, column=1, value=list_data[0]) #A1
        # self.sheet.cell(row=1, column=2, value=list_data[1]) #B1
        # self.sheet.cell(row=1, column=3, value=list_data[2]) #C1
    def save_file(self):
        self.workbook.save(self.save_path)
    def merge_cell(self,row_or_col,target,start,stop):
        if row_or_col == "row" or row_or_col == "Row":
            self.sheet.merge_cells(start_row=target, start_column=start, end_row=target, end_column=stop)
        elif row_or_col == "col" or row_or_col == "column" or row_or_col == "Column" or row_or_col == "Col":
            self.sheet.merge_cells(start_row=start, start_column=target, end_row=stop, end_column=target)
        else:
            print(None)
    def make_boder(self,start_row,end_row,start_col,end_col):#ทำเส้นล้อมตาราง
        border_style = Border(
            left=Side(border_style='thin', color='000000'),  # Thin black border on the left
            right=Side(border_style='thin', color='000000'),  # Thin black border on the right
            top=Side(border_style='thin', color='000000'),  # Thin black border on the top
            bottom=Side(border_style='thin', color='000000')  # Thin black border on the bottom
        )
        for row in self.sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            for cell in row:
                cell.border = border_style        

class header(WriteExcel):
    def __init__(self,save_path):
        self.save_path = save_path
        super().__init__(self.save_path)
        super().create_workbook()
        super().select_sheet("CPE")
        
        super().merge_cell("row",1,1,27)
        super().insert_data(1,1,"ตารางเรียนคณะวิศวกรรมศาสตร์ ศรีราชา มหาวิทยาลัยเกษตรศาสตร์ วิทยาเขตศรีราชา","all_center")
        super().merge_cell("row",2,1,27)
        super().insert_data(2,1,"ประจำภาคต้น ปีการศึกษา 2566","all_center")

        super().insert_data(4,4,"บรรยาย","center,bottom","bold")
        super().merge_cell("row",4,4,14)
        super().insert_data(4,15,"ปฏิบัติ","center,bottom","bold")
        super().merge_cell("row",4,15,24)

        super().insert_data(4,1,"รหัสวิชา","center,bottom","bold")
        super().merge_cell("col",1,4,5)
        super().insert_data(4,2,"รหัสวิชา-พ.ศ.หลักสูตร","center,bottom","bold")
        super().merge_cell("col",2,4,5)
        super().insert_data(4,3,"ชื่อวิชา","center,bottom","bold")
        super().merge_cell("col",3,4,5)

        Sub_Head_Data = ["หน่วยกิต","หน่วย","จำนวน ชม.","หมู่","วัน","เริ่ม","-","สิ้นสุด","ห้อง","สาขา","จำนวน"]
        count = 0
        for i in range(4,25):
            if count == len(Sub_Head_Data):
                count = 1
                super().insert_data(5,i,Sub_Head_Data[count],"center,bottom","bold")
                count+=1
            else:
                super().insert_data(5,i,Sub_Head_Data[count],"center,bottom","bold")
                count+=1
                
        super().insert_data(4,25,"อาจารย์","center,bottom","bold")
        super().merge_cell("col",25,4,5)
        super().insert_data(4,26,"CODE-อาจารย์-1","center,bottom","bold")
        super().merge_cell("col",26,4,5)
        super().insert_data(4,27,"วิชาพื้นฐาน","center,bottom","bold")
        super().merge_cell("col",27,4,5)

        super().insert_data(6,1,"กลุ่มวิชาวิศวกรรมคอมพิวเตอร์",None,None)
        super().merge_cell("row",6,1,27)

        super().insert_data(7,1,"สาขาวิชาวิศวกรรมคอมพิวเตอร์และสารสนเทศศาสตร์ (T12)",None,None)
        super().merge_cell("row",7,1,27)
        
        #super().save_file()